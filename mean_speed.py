import numpy as np
import argparse
import cv2
import pandas as pd
import math
import matplotlib.pyplot as plt
import openpyxl 
from openpyxl.styles import PatternFill
#https://www.codespeedy.com/how-to-add-color-to-excel-cells-using-python/ 
# для того чтобы цвета ячеек эксель менять
def add_colors(colors, track_path):
    path = get_pathes(track_path, 'E')
    wb = openpyxl.load_workbook(path)
    ws = wb['speed']
    for no, c in enumerate(colors):
        fill_cell = PatternFill(patternType='solid', fgColor=c)
        cell = 'B' + str(no + 2)
        ws[cell].fill = fill_cell
    wb.save(path)
    
def rgb_to_hex(r, g, b):
    return '{:02x}{:02x}{:02x}'.format(r, g, b).upper()

def read_tracks_from_txt(path):
    frames = []
    inds = []
    color = []
    tracks = []
    last_no = 0
    with open(path) as f:
        for i in f:
            dic = {}
            a = list(map(str, i[:-2].split(' ')))
            if (len(a) - 4) % 5 != 0:
                print(len(a) - 6)
            else:
                no = int(a[0])
                frame_ind = int(a[1])
                identifier = int(a[2])
                #r = int(a[3])
                #g = int(a[4])
                #b = int(a[5])
                #c = rgb_to_hex(r, g, b)
                c = a[3]
                a = np.array([float(x) for x in a[4:]]).reshape((-1, 5)).tolist()
                #dic[frame_ind] = a
                #tracks.append(dic)
                
                frames.append(frame_ind)
                inds.append(identifier)
                color.append(c)
                tracks.append(a)
            
    return [inds, frames, color, tracks]
 
 
def mean_speed_for_ant(traj, dt):
    first_p = traj[0]
    traj = traj[1:]
    all_speed = []
    len_of_traj = 0
    for point in traj:
        l = math.dist(first_p, point)
        len_of_traj += len_of_traj
        speed = l/dt
        all_speed.append(speed)
        first_p = point
    #print(f"Скорость по особям: {sum(all_speed)/len(all_speed)}")
    return all_speed, sum(all_speed)/len(all_speed)

def plot_gist(d_mean_ants, d_mean_steps, d_ns_mean_ants, d_ns_mean_steps, path):
    lenth = max([len(d_mean_ants), len(d_mean_steps), len(d_ns_mean_ants), len(d_ns_mean_steps)])
    none_array = [None] * lenth
    #дополни массивы none тогда они будут одной длины, или попробуй список массивов
    data = [np.array(d_mean_ants), np.array(d_mean_steps), np.array(d_ns_mean_ants), np.array(d_ns_mean_steps)]
    fig, ax = plt.subplots()
    ax.set_title('Распределение скоростей')
    parts = ax.violinplot(
        data, showmeans=False, showmedians=False,
        showextrema=False)
    
    for pc in parts['bodies']:
        pc.set_facecolor('#D43F3A')
        pc.set_edgecolor('black')
        pc.set_alpha(1)
    
    #medians = np.percentile(data, [50], axis=1)
    medians = [np.percentile(d_mean_ants, [50]), np.percentile(d_mean_steps, [50]), np.percentile(d_ns_mean_ants, [50]), np.percentile(d_ns_mean_steps, [50])]

    ax.scatter([1, 2, 3, 4], medians, marker='o', color='white', s=30, zorder=3)
    
    labels = ['Ост./особи', 'Ост./шаги', 'Б.Ост/особи', 'Б.Ост/шаги']
    
    ax.xaxis.set_tick_params(direction='out')
    ax.xaxis.set_ticks_position('bottom')
    ax.set_xticks([1, 2, 3, 4])
    ax.set_xticklabels(labels)
    ax.set_xlim(0.25, len(labels) + 0.75)
    ax.set_xlabel('Sample name')
    plt.savefig(path)
    
    
def lenth_of_traj(traj):
    first_p = traj[0]
    traj = traj[1:]
    len_of_traj = 0
    for point in traj:
        l = math.dist(first_p, point)
        len_of_traj += l
        first_p = point
    return len_of_traj


def get_pathes(path, param):
    folder = path[:path.rfind('/')]
    name = path[path.rfind('/'):path.rfind('_')]
    if param == 'V':
        return folder + name + '.mp4'
    if param == 'C':
        return folder + name + '_speed.csv'
    if param == 'G':
        return folder + name + '_distrib.png'
    if param == 'E':
        return folder + name + '_speed.xlsx'
        
def get_speed(p_1, p_2, dt):
    l = math.dist(p_1, p_2)
    speed = l/dt
    return speed


def count_mean_speed(track_path):
    #Все траектории
    all_ants = read_tracks_from_txt(track_path)
    video_path = get_pathes(track_path, 'V')
    cap = cv2.VideoCapture(video_path)
    fps = cap.get(cv2.CAP_PROP_FPS)
    cap.release()
    dt = 1/fps
    #Средняя скорость треков с остановок
    d_mean_ants = []
    d_mean_steps = []
    #Средняя скорость треков без остановок
    d_ns_mean_ants = []
    d_ns_mean_steps = []
    mean_speed = []
    #Длины каждой траектории
    lenthes = []
    #Время остановок
    time_of_stops = []
    #Время движения
    time_of_movement = []
    #Время начала отслеживания
    time_of_start = []
    v_min = 0.005
    mean_speed_by_dist = []
    for i in range(len(all_ants[0])):
        
        ind = all_ants[0][i]
        frame = all_ants[1][i]
        color = all_ants[2][i]
        track = all_ants[3][i]
        #frame = list(ant.keys())[0]
        #track = list(ant.values())[0]
        time_of_start.append(round(frame * dt, 1))
        curr_pts = []
        mean_ants = []
        speed = 0
        lenth = len(track)
        frames_dont_move = 0
        for num, tr in enumerate(track):
            curr_v = 0
            if num != 0:
                curr_v = get_speed([tr[0], tr[1]], [track[num-1][0], track[num-1][1]], dt)
            #print(f"{tr[3]} : {curr_v}")
            if curr_v > v_min:
                speed += tr[3]
                curr_pts.append([tr[0], tr[1]])
            else:
                frames_dont_move += 1
            mean_ants.append([tr[0], tr[1]])#для подсчета распределения скоростей с остановками
        # в момент появления муравья его скорость = 0, чтобы не учитывать это
        time_of_stops.append(frames_dont_move - 1)
        time_of_movement.append(len(track)-frames_dont_move + 1)
        
        ant_speeds, mean_ant_speed = mean_speed_for_ant(mean_ants, dt)
        lenthes.append(round(lenth_of_traj(mean_ants), 4))
        ns_ant_speeds, ns_mean_ant_speed = mean_speed_for_ant(curr_pts, dt)
            
        d_ns_mean_ants.append(round(ns_mean_ant_speed, 4))
        d_mean_ants.append(round(mean_ant_speed, 4))
            
        d_mean_steps += ant_speeds
        d_ns_mean_steps += ns_ant_speeds
        if speed != 0:
            mean_speed.append(speed/lenth)
            
    time_of_stops = (round(x * dt, 4) for x in time_of_stops)
    time_of_movement = (round(x * dt, 4) for x in time_of_movement)
    
    df = pd.DataFrame({'Муравей': all_ants[0], 'Цвет': all_ants[2], 'Время начала отслеживания (с)': time_of_start,'Длина пути (м)': lenthes, 'Время остановок (с)': time_of_stops, 'Время пути (с)': time_of_movement, 'Средняя скорость с остановками (м/c)': d_mean_ants, 'Средняя скорость без остановок (м/c)': d_ns_mean_ants})
    sort_df = df.sort_values(by ='Муравей')
    #save_path = get_pathes(track_path, 'C')
    #df.to_csv(save_path, index=False)
    save_path = get_pathes(track_path, 'E')
    sort_df.to_excel(save_path, sheet_name="speed",index=False)
    add_colors(sort_df['Цвет'].to_list(), track_path)
    gist_path = get_pathes(track_path, 'G')
    #plot_gist(d_mean_ants, d_mean_steps, d_ns_mean_ants, d_ns_mean_steps, gist_path)
    #if mean_speed:
    #    print(f"Средняя скорость по Калману: {sum(mean_speed)/len(mean_speed)} м/c")
    print(f"Средняя скорость по дистанции: {sum(d_ns_mean_ants)/len(d_ns_mean_ants)} м/c")


if __name__ == '__main__':
    # Нужно реализровать вывод таблицы: для каждого трека: длина пути, время остановок, время пути, средняя скорость без остановок, средняя скорость с остановками.
    parser = argparse.ArgumentParser()
    parser.add_argument('--tracks_path', nargs='?', default="/home/ubuntu/ant_detection/problems/full_video/18.08.20_Fp2_плос2_tracks.txt", help="Specify yaml track path", type=str)
    args = parser.parse_args()
    
    print('**********************НАЧАЛО ОПРЕДЕЛЕНИЯ СРЕДНИХ СКОРОСТЕЙ********************')
    count_mean_speed(args.tracks_path)
    print('***********************КОНЕЦ ОПРЕДЕЛЕНИЯ СРЕДНИХ СКОРОСТЕЙ********************')
